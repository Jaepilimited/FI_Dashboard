# -*- coding: utf-8 -*-
"""손익계산서 엑셀 내보내기 (대시보드 구성·색상 동일, 원 단위)

생성 파일 (기본: ~/Downloads):
  손익계산서_지역별_<기간>.xlsx     대륙(Continent1) 시트 12개 × [월별 국가 Top10+기타+합계]
  손익계산서_상품별_<기간>.xlsx     라인(Line) 시트 × [월별 SKU Top10+기타+합계]
  손익계산서_판매유형별_<기간>.xlsx  1시트 × [월별 B2B/B2C/기타+합계]

열 순위(국가/SKU)는 1~5월 누계 매출 기준으로 고정.
사용: python export_pl_excel.py [--months 2026-01 ... ] [--out DIR]
"""
import argparse
import os
import sys

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

BASE = "http://127.0.0.1:5000"
LOGIN = {"username": "jeffrey", "password": "skin1004!"}
DEFAULT_MONTHS = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05"]
TOP_N = 10

# 대시보드 CPL_ROW_DEFS와 동일한 행 구성
ROWS = [
    ("sales",    "매출액",                 "bold", None),
    ("cogs",     "매출원가",               "",     None),
    ("gross",    "매출총이익",             "bold", None),
    ("sgaD",     "판매관리비(직접)",       "bold", None),
    ("sgaD_adv", "광고선전비",             "sub",  None),
    ("sgaD_log", "물류비",                 "sub",  None),
    ("sgaD_fee", "수수료",                 "sub",  None),
    ("sgaD_hr",  "인건비",                 "sub",  None),
    ("sgaD_etc", "기타",                   "sub",  None),
    ("direct",   "직접이익",               "bold", "direct"),
    ("sgaO",     "판매관리비 간접비(조직)", "bold", None),
    ("sgaO_adv", "광고선전비",             "sub",  None),
    ("sgaO_log", "물류비",                 "sub",  None),
    ("sgaO_fee", "수수료",                 "sub",  None),
    ("sgaO_hr",  "인건비",                 "sub",  None),
    ("sgaO_etc", "기타",                   "sub",  None),
    ("contrib",  "공헌이익",               "bold", "contrib"),
    ("sgaC",     "판매관리비 간접비(전사)", "bold", None),
    ("sgaC_adv", "광고선전비",             "sub",  None),
    ("sgaC_log", "물류비",                 "sub",  None),
    ("sgaC_fee", "수수료",                 "sub",  None),
    ("sgaC_hr",  "인건비",                 "sub",  None),
    ("sgaC_etc", "기타",                   "sub",  None),
    ("op",       "영업이익",               "bold", "op"),
]
BASE_KEYS = [
    "sales", "cogs", "gross",
    "sgaD", "sgaD_adv", "sgaD_log", "sgaD_fee", "sgaD_hr", "sgaD_etc",
    "sgaO", "sgaO_adv", "sgaO_log", "sgaO_fee", "sgaO_hr", "sgaO_etc",
    "sgaC", "sgaC_adv", "sgaC_log", "sgaC_fee", "sgaC_hr", "sgaC_etc",
]

# 대시보드 라이트 테마 색 (rgba 틴트를 흰 배경에 합성한 값)
HEADER_BG = "FF4B5568"
HL_BG = {"direct": "FFEDF9F1", "contrib": "FFEBF5FF", "op": "FFFFEFEF"}
PCT_HL_BG = {"direct": "FFF6FCF8", "contrib": "FFF4F9FF", "op": "FFFFF7F7"}
SUB_BG = "FFEEF0F4"
SUB_PCT_BG = "FFF5F6F9"
TEXT = "FF0F172A"
TEXT_SUB = "FF64748B"
TEXT_PCT = "FF94A3B8"
NEG = "FFE11D48"
THIN = Side(style="thin", color="FFD9D9D9")
STRONG = Side(style="medium", color="FFB9BEC9")
BORDER = Border(top=THIN, left=THIN, bottom=THIN, right=THIN)


def api(sess, path, **params):
    r = sess.get(BASE + path, params=params)
    r.raise_for_status()
    return r.json()


def node_series(node, resp_months, months):
    """API 응답 node → {key: [월별 값(원)]}, 파생계정 포함."""
    idx = {m: i for i, m in enumerate(resp_months)}
    out = {}
    for k in BASE_KEYS:
        arr = node.get(k) or [0] * len(resp_months)
        out[k] = [arr[idx[m]] if m in idx else 0 for m in months]
    out["direct"] = [out["gross"][i] - out["sgaD"][i] for i in range(len(months))]
    out["contrib"] = [out["direct"][i] - out["sgaO"][i] for i in range(len(months))]
    out["op"] = [out["contrib"][i] - out["sgaC"][i] for i in range(len(months))]
    return out


def add_series(a, b):
    for k in a:
        a[k] = [a[k][i] + b[k][i] for i in range(len(a[k]))]


def zero_series(n):
    keys = BASE_KEYS + ["direct", "contrib", "op"]
    return {k: [0] * n for k in keys}


def build_columns(resp, months, top_n=TOP_N):
    """응답 nodes → [(열이름, series)] : 누계 매출 상위 top_n + 기타 + 합계."""
    pairs = [(n["name"], node_series(n, resp["months"], months)) for n in resp["nodes"]]
    return columns_from_pairs(pairs, months, top_n)


def columns_from_pairs(pairs, months, top_n, sort=True):
    nodes = list(pairs)
    if sort:
        nodes.sort(key=lambda kv: -sum(kv[1]["sales"]))
    top = nodes[:top_n]
    rest = nodes[top_n:]
    cols = list(top)
    if rest:
        etc = zero_series(len(months))
        for _, s in rest:
            add_series(etc, s)
        cols.append(("기타", etc))
    total = zero_series(len(months))
    for _, s in nodes:
        add_series(total, s)
    cols.append(("합계", total))
    return cols


def sheet_name(raw, used):
    name = str(raw)
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    name = name.strip()[:31] or "시트"
    base, i = name, 2
    while name in used:
        name = f"{base[:28]}_{i}"
        i += 1
    used.add(name)
    return name


def write_sheet(wb, title, months, cols, used):
    ws = wb.create_sheet(sheet_name(title, used))
    n_per = len(cols)
    header_font = Font(size=10, bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 헤더 2행: 월 그룹 / 열 이름
    ws.cell(1, 1, "구분")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    for mi, m in enumerate(months):
        c0 = 2 + mi * n_per
        ws.cell(1, c0, m.replace("-", "."))
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + n_per - 1)
        for ci, (name, _) in enumerate(cols):
            ws.cell(2, c0 + ci, name)
    for r in (1, 2):
        for c in range(1, 2 + len(months) * n_per):
            cell = ws.cell(r, c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = BORDER

    # 데이터 행: 계정 + % 행
    r = 3
    for key, label, kind, hl in ROWS:
        bg = HL_BG.get(hl) or (SUB_BG if kind == "sub" else None)
        font = Font(size=10, bold=(kind == "bold" or hl is not None),
                    color=TEXT_SUB if kind == "sub" else TEXT)
        lab = ws.cell(r, 1, ("    " + label) if kind == "sub" else label)
        lab.font = font
        lab.alignment = left
        lab.border = BORDER
        if bg:
            lab.fill = PatternFill("solid", fgColor=bg)
        for mi in range(len(months)):
            for ci, (_, s) in enumerate(cols):
                v = int(round(s[key][mi]))
                cell = ws.cell(r, 2 + mi * n_per + ci, v)
                cell.number_format = "#,##0"
                is_total = ci == n_per - 1
                cell.font = Font(size=10, bold=(kind == "bold" or hl is not None or is_total),
                                 color=NEG if v < 0 else (TEXT_SUB if kind == "sub" else TEXT))
                cell.alignment = right
                cell.border = Border(top=THIN, bottom=THIN, right=THIN,
                                     left=STRONG if ci == 0 else THIN)
                if bg:
                    cell.fill = PatternFill("solid", fgColor=bg)
        r += 1
        # % 행 (매출 대비)
        pbg = PCT_HL_BG.get(hl) or (SUB_PCT_BG if kind == "sub" else None)
        plab = ws.cell(r, 1, "      %")
        plab.font = Font(size=8, color=TEXT_PCT)
        plab.alignment = left
        plab.border = BORDER
        if pbg:
            plab.fill = PatternFill("solid", fgColor=pbg)
        for mi in range(len(months)):
            for ci, (_, s) in enumerate(cols):
                sales = s["sales"][mi]
                cell = ws.cell(r, 2 + mi * n_per + ci,
                               (s[key][mi] / sales) if sales else None)
                cell.number_format = "0.0%"
                cell.font = Font(size=8, color=TEXT_PCT)
                cell.alignment = right
                cell.border = Border(top=THIN, bottom=THIN, right=THIN,
                                     left=STRONG if ci == 0 else THIN)
                if pbg:
                    cell.fill = PatternFill("solid", fgColor=pbg)
        r += 1

    ws.column_dimensions["A"].width = 24
    for c in range(2, 2 + len(months) * n_per):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B3"
    return ws


def export_region(sess, months, out_dir, tag):
    resp1 = api(sess, "/api/pl", dim="Continent1", months=months)
    continents = sorted(resp1["nodes"], key=lambda n: -sum(n["sales"]))
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for node in continents:
        cont = node["name"]
        resp = api(sess, "/api/pl", dim="Country", months=months, continent1=cont)
        if not resp["nodes"]:
            continue
        cols = build_columns(resp, months)
        write_sheet(wb, cont, months, cols, used)
        print(f"  [지역별] {cont}: 국가 {len(resp['nodes'])}개 → 열 {len(cols)}")
    path = os.path.join(out_dir, f"손익계산서_지역별_{tag}.xlsx")
    wb.save(path)
    return path


def export_product(sess, months, out_dir, tag):
    resp1 = api(sess, "/api/pl", dim="Line", months=months)
    lines = [n["name"] for n in sorted(resp1["nodes"], key=lambda n: -sum(n["sales"]))
             if n["name"] not in ("Others", "조정")]
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for line in lines:
        resp = api(sess, "/api/pl", dim="Product_Name", months=months, line=line)
        if not resp["nodes"]:
            continue
        cols = build_columns(resp, months)
        write_sheet(wb, line, months, cols, used)
        print(f"  [상품별] {line}: SKU {len(resp['nodes'])}개 → 열 {len(cols)}")
    path = os.path.join(out_dir, f"손익계산서_상품별_{tag}.xlsx")
    wb.save(path)
    return path


def export_sales_type(sess, months, out_dir, tag):
    """시트: SK / 유통(brand=UM) / 기타(전체−SK−유통). 열: 월별 B2B/B2C/기타 + 합계."""
    order = {"B2B": 0, "B2C": 1}
    n = len(months)

    def series_map(resp):
        return {nd["name"]: node_series(nd, resp["months"], months) for nd in resp["nodes"]}

    m_all = series_map(api(sess, "/api/pl", dim="Sales_Type", months=months))
    m_sk = series_map(api(sess, "/api/pl", dim="Sales_Type", months=months, brand="SK"))
    m_um = series_map(api(sess, "/api/pl", dim="Sales_Type", months=months, brand="UM"))

    names = sorted(set(m_all) | set(m_sk) | set(m_um), key=lambda x: order.get(x, 9))
    m_etc = {}
    for name in names:
        s = {k: list(v) for k, v in (m_all.get(name) or zero_series(n)).items()}
        for part in (m_sk, m_um):
            if name in part:
                for k in s:
                    s[k] = [s[k][i] - part[name][k][i] for i in range(n)]
        m_etc[name] = s

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for title, mp in (("SK", m_sk), ("유통", m_um), ("기타", m_etc)):
        pairs = [(name, mp[name]) for name in names if name in mp]
        cols = columns_from_pairs(pairs, months, top_n=len(pairs), sort=False)
        write_sheet(wb, title, months, cols, used)
        print(f"  [판매유형별] {title}: 열 {len(cols)}")
    path = os.path.join(out_dir, f"손익계산서_판매유형별_{tag}.xlsx")
    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--months", nargs="+", default=DEFAULT_MONTHS)
    ap.add_argument("--out", default=os.path.join(os.path.expanduser("~"), "Downloads"))
    ap.add_argument("--only", choices=["region", "product", "sales"], default=None)
    args = ap.parse_args()
    months = sorted(args.months)
    tag = f"{months[0]}~{months[-1]}"

    sess = requests.Session()
    r = sess.post(BASE + "/login", data=LOGIN)
    if "/login" in r.url:
        sys.exit("로그인 실패 — 서버/계정 확인")

    targets = {"region": export_region, "product": export_product, "sales": export_sales_type}
    fns = [targets[args.only]] if args.only else list(targets.values())
    for fn in fns:
        path = fn(sess, months, args.out, tag)
        print("저장:", path)


if __name__ == "__main__":
    main()
